#!/usr/bin/env python3
"""Generate spot card master spreadsheet from resources/spot_cards plus design suggestions."""

from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
SPOT_CARDS = ROOT / "resources" / "spot_cards"
OUT_CSV = ROOT / "docs" / "spot_cards_master_spreadsheet.csv"
OUT_XLSX = ROOT / "docs" / "spot_cards_master_spreadsheet.xlsx"

HEADERS = [
    "Status",
    "Shelf",
    "Rarity",
    "Name",
    "Id",
    "Product_Or_Sigil",
    "Restriction",
    "Effect",
    "Primary_Identity",
    "Secondary_Identities",
    "Starter_Eligible",
    "Layout_Gate",
    "In_First_Cut",
    "Notes",
]

RARITY_NAMES = {0: "Common", 1: "Uncommon", 2: "Rare"}
SHELF_BY_FOLDER = {
    "ingredients": "Ingredients",
    "kitchenware": "Kitchenware",
    "economy": "Economy",
    "utilities": "Utility",
}
SHELF_ORDER = {"Ingredients": 0, "Kitchenware": 1, "Economy": 2, "Utility": 3}
RARITY_ORDER = {"Common": 0, "Uncommon": 1, "Rare": 2}
STATUS_ORDER = {"Existing": 0, "Suggested": 1, "Removed": 2}

STARTER_IDS = {"tomato", "salt"}

# Design backlog. Feast copy only. Not in spot_cards until implemented.
SUGGESTED = [
    (
        "Suggested",
        "Ingredients",
        "Common",
        "Capacitor",
        "capacitor",
        "Flavour",
        "",
        "+16 Flavour. Store 10 charge on this card.",
        "Charge",
        "",
        "No",
        "",
        "Yes",
        "Tomato is +20. Capacitor banks charge for Release / Overfill.",
    ),
    (
        "Suggested",
        "Kitchenware",
        "Common",
        "Trickle Charge",
        "trickle_charge",
        "Charge",
        "",
        "Add 12 charge to the next Ingredient on this course.",
        "Charge",
        "",
        "No",
        "",
        "Yes",
        "Feeds the Charge line without being a weak Ingredient.",
    ),
    (
        "Suggested",
        "Ingredients",
        "Uncommon",
        "Release",
        "release",
        "Flavour",
        "",
        "Gain Flavour equal to this card's charge, then clear charge. If you dumped 20 or more, also +2 Mult.",
        "Charge",
        "",
        "No",
        "",
        "Yes",
        "Empty brick. Charged it can beat Tomato.",
    ),
    (
        "Suggested",
        "Ingredients",
        "Rare",
        "Overfill",
        "overfill",
        "Flavour",
        "",
        "When charge reaches 30, dump all charge as Flavour and Double this card.",
        "Charge",
        "Double",
        "No",
        "",
        "Yes",
        "Rare Charge spike.",
    ),
    (
        "Suggested",
        "Economy",
        "Common",
        "Mint Cell",
        "mint_cell",
        "Gold",
        "",
        "+2 Gold.",
        "Gold Ledger",
        "",
        "No",
        "",
        "Yes",
        "Tip is +1. Common Economy producer gap.",
    ),
    (
        "Suggested",
        "Ingredients",
        "Uncommon",
        "Vault Mult",
        "vault_mult",
        "Mult",
        "",
        "+2 Mult, plus +1 Mult per 8 Gold you hold.",
        "Gold Ledger",
        "",
        "No",
        "",
        "Yes",
        "Cover Charge already converts Gold to Flavour.",
    ),
    (
        "Suggested",
        "Kitchenware",
        "Uncommon",
        "Reserve Double",
        "reserve_double",
        "Double",
        "",
        "If you hold at least 8 Gold, Double the next Ingredient. Does not spend Gold.",
        "Gold Ledger",
        "Double",
        "No",
        "",
        "Yes",
        "Replaces spend-on-board double tools.",
    ),
    (
        "Suggested",
        "Ingredients",
        "Rare",
        "Peak Hit",
        "peak_hit",
        "Flavour",
        "",
        "+24 Flavour. If Doubled, also +6 Mult.",
        "Double",
        "",
        "No",
        "",
        "Yes",
        "Lemonade is the small Doubled payoff.",
    ),
    (
        "Suggested",
        "Ingredients",
        "Rare",
        "Last Claim",
        "last_claim",
        "Flavour",
        "",
        "+55 Flavour if a card already spoiled on this course this hour, else +16.",
        "Spoil (tool)",
        "",
        "No",
        "",
        "Yes",
        "Miss below Tomato on purpose. Hit is huge.",
    ),
    (
        "Suggested",
        "Ingredients",
        "Uncommon",
        "Opening Ratio",
        "opening_ratio",
        "Mult",
        "",
        "+6 Mult if first Ingredient in course, else +2.",
        "Bookend",
        "",
        "No",
        "",
        "Yes",
        "Bookend Mult pair with Last Ratio.",
    ),
    (
        "Suggested",
        "Ingredients",
        "Uncommon",
        "Last Ratio",
        "last_ratio",
        "Mult",
        "",
        "+6 Mult if last Ingredient in course, else +2.",
        "Bookend",
        "",
        "No",
        "",
        "Yes",
        "Bookend Mult pair with Opening Ratio.",
    ),
    (
        "Suggested",
        "Ingredients",
        "Rare",
        "Lone Ratio",
        "lone_ratio",
        "Mult",
        "One-spot course",
        "+8 Mult. 1-spot course only.",
        "Solo Cell",
        "",
        "No",
        "Size-1 course",
        "Yes",
        "Steak is +80 Flavour on the same seat.",
    ),
    (
        "Suggested",
        "Ingredients",
        "Uncommon",
        "Mult Sink",
        "mult_sink",
        "Mult",
        "",
        "+3 Mult per course that already received a Pass this hour.",
        "Course Pass",
        "",
        "No",
        "",
        "Yes",
        "Cumin is flat +5 Mult to the next course.",
    ),
    (
        "Suggested",
        "Ingredients",
        "Uncommon",
        "Pair Mult",
        "pair_mult",
        "Mult",
        "",
        "+3 Mult per adjacent Next Seasoning.",
        "Next cluster",
        "",
        "No",
        "",
        "Yes",
        "Corn is the Flavour adjacency card.",
    ),
    (
        "Suggested",
        "Kitchenware",
        "Common",
        "Gap Stamp",
        "gap_stamp",
        "Proof",
        "",
        "If an adjacent Next spot is empty, the next Ingredient permanently +10 Flavour.",
        "Sparse",
        "Proof",
        "No",
        "",
        "Yes",
        "Oven Mitt is +5 to lowest Next with no gap requirement.",
    ),
]

REMOVED = [
    (
        "Removed",
        "Ingredients",
        "Uncommon",
        "Golden Ratio",
        "golden_ratio",
        "Mult",
        "",
        "+3 Mult. Gain +1 Mult for each Gold spent this day.",
        "Gold Ledger",
        "",
        "No",
        "",
        "No",
        "Removed from resources. Do not restore.",
    ),
]


def _read_field(text: str, field: str) -> str:
    match = re.search(rf'^{re.escape(field)} = (?:"([^"]*)"|(\d+))', text, flags=re.MULTILINE)
    if not match:
        return ""
    return match.group(1) if match.group(1) is not None else match.group(2)


def _shelf_from_path(path: Path) -> str:
    for part in path.parts:
        if part in SHELF_BY_FOLDER:
            return SHELF_BY_FOLDER[part]
    return "Ingredients"


def _product_label(raw_type: str, description: str) -> str:
    mapping = {
        "0": "Flavour",
        "1": "Mult",
        "2": "Gold",
        "3": "Hybrid",
    }
    if raw_type in mapping:
        return mapping[raw_type]
    desc = description.lower()
    if "double" in desc or "doubled" in desc:
        return "Double"
    if "fire again" in desc or "fires again" in desc:
        return "Again"
    if "pass " in desc or "passes" in desc:
        return "Pass"
    if "proof" in desc or "permanently" in desc:
        return "Proof"
    if "gold" in desc and "flavour" not in desc and "mult" not in desc:
        return "Gold"
    if "mult" in desc and "flavour" not in desc:
        return "Mult"
    if "flavour" in desc:
        return "Flavour"
    return ""


def scan_spot_cards() -> list[tuple]:
    rows: list[tuple] = []
    for path in sorted(SPOT_CARDS.rglob("*.tres")):
        text = path.read_text(encoding="utf-8")
        card_id = _read_field(text, "id")
        if not card_id:
            continue
        name = _read_field(text, "name")
        description = _read_field(text, "description")
        rarity_raw = _read_field(text, "rarity")
        rarity = RARITY_NAMES.get(int(rarity_raw), "Common") if rarity_raw.isdigit() else "Common"
        product_type = _read_field(text, "type")
        shelf = _shelf_from_path(path)
        product = _product_label(product_type, description)
        starter = "Yes" if card_id in STARTER_IDS else "No"
        rows.append(
            (
                "Existing",
                shelf,
                rarity,
                name,
                card_id,
                product,
                "",
                description,
                "",
                "",
                starter,
                "",
                "No",
                f"Loaded from {path.relative_to(ROOT).as_posix()}",
            )
        )
    return rows


def sort_key(row: tuple) -> tuple:
    return (
        SHELF_ORDER.get(row[1], 9),
        RARITY_ORDER.get(row[2], 9),
        STATUS_ORDER.get(row[0], 9),
        row[3].lower(),
    )


def build_summary_rows(rows: list[tuple]) -> list[list]:
    summary: list[list] = []
    existing = [r for r in rows if r[0] == "Existing"]
    suggested = [r for r in rows if r[0] == "Suggested"]
    removed = [r for r in rows if r[0] == "Removed"]
    first_cut = [r for r in rows if r[12] == "Yes"]

    summary.append(["=== SUMMARY ==="])
    summary.append([])
    summary.append(["Metric", "Count"])
    summary.append(["Total cards in spreadsheet", len(rows)])
    summary.append(["Existing (in spot_cards)", len(existing)])
    summary.append(["Suggested (design backlog)", len(suggested)])
    summary.append(["Removed from repo", len(removed)])
    summary.append(["Suggested first-cut bundle", len(first_cut)])
    summary.append([])

    summary.append(["By shelf (Existing)"])
    shelf_counts = Counter(r[1] for r in existing)
    for shelf, count in sorted(shelf_counts.items(), key=lambda item: SHELF_ORDER.get(item[0], 9)):
        summary.append([shelf, count])
    summary.append([])

    summary.append(["By rarity (Existing)"])
    for rarity in ("Common", "Uncommon", "Rare"):
        count = sum(1 for r in existing if r[2] == rarity)
        summary.append([rarity, count])
    summary.append([])

    summary.append(["Terminology"])
    summary.append(["Player formula", "Flavour x Mult = Rating"])
    summary.append(["Day / Hour", "round / turn in code"])
    summary.append(["Course / Spot", "segment / tile in code"])
    summary.append(["Events catalog", "docs/events_catalog.csv"])
    summary.append(["Glossary", "docs/feast_glossary.md"])
    return summary


def write_csv(rows: list[tuple], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(rows)
        writer.writerow([])
        for row in build_summary_rows(rows):
            writer.writerow(row)


def write_xlsx(rows: list[tuple], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    suggested_fill = PatternFill(fill_type="solid", fgColor="FFF9C4")
    removed_fill = PatternFill(fill_type="solid", fgColor="E0E0E0")

    wb = Workbook()
    wb.remove(wb.active)

    def write_sheet(title: str, sheet_rows: list[tuple]) -> None:
        ws = wb.create_sheet(title)
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = "A2"
        for row_idx, row_data in enumerate(sheet_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            fill = None
            if row_data[0] == "Suggested":
                fill = suggested_fill
            elif row_data[0] == "Removed":
                fill = removed_fill
            if fill:
                for col_idx in range(1, len(HEADERS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    write_sheet("All Cards", rows)
    for shelf in ("Ingredients", "Kitchenware", "Economy", "Utility"):
        write_sheet(shelf, [r for r in rows if r[1] == shelf])

    summary_sheet = wb.create_sheet("Summary")
    for row_idx, row_data in enumerate(build_summary_rows(rows), start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
            if row_data and str(row_data[0]).startswith("==="):
                cell.font = Font(bold=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    rows = sorted(scan_spot_cards() + SUGGESTED + REMOVED, key=sort_key)
    write_csv(rows, OUT_CSV)
    print(f"Wrote {len(rows)} rows to {OUT_CSV}")
    try:
        write_xlsx(rows, OUT_XLSX)
        print(f"Wrote workbook to {OUT_XLSX}")
    except ImportError:
        print("openpyxl not installed. CSV written only.")
        print("Install with: pip install -r scripts/debug/requirements.txt")


if __name__ == "__main__":
    main()
